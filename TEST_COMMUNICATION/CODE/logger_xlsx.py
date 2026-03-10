from datetime import datetime
from pathlib import Path
import time
import requests
from openpyxl import Workbook, load_workbook

URL = "http://192.168.4.1/api/state"
PERIOD_S = 1.0

filename = Path(f"bracelet_raw_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "raw_data"
ws.append(["timestamp", "tempC", "humPct", "adcRaw", "led", "touch", "buzzer"])
wb.save(filename)

print("Logging to", filename)

try:
    while True:
        try:
            r = requests.get(URL, timeout=2)
            r.raise_for_status()
            s = r.json()

            wb = load_workbook(filename)
            ws = wb["raw_data"]
            ws.append([
                datetime.now().isoformat(timespec="seconds"),
                s.get("tempC"),
                s.get("humPct"),
                s.get("adcRaw"),
                s.get("led"),
                s.get("touch"),
                s.get("buzzer"),
            ])
            wb.save(filename)
            print(".", end="", flush=True)
        except Exception as e:
            print("\nFetch error:", e, flush=True)

        time.sleep(PERIOD_S)

except KeyboardInterrupt:
    print("\nStopped. File saved:", filename)
